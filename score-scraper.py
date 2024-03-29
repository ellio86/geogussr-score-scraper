from openpyxl import load_workbook, Workbook
import openpyxl
from dataclasses import dataclass, field
from selenium import webdriver
import datetime
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import sys

# Check for arguement from cmdline
try:
    FILENAME = sys.argv[1]
except IndexError:
    FILENAME = "Geoguessr.xlsx"
    
# Open workbook and set current sheet
try:
    workbook = load_workbook(filename=FILENAME)
    sheet = workbook.active
except FileNotFoundError:
    c=input(f"{FILENAME} not found! would you like to create it? Y/n ")
    if input == "n":
        quit()
    else:
        workbook = Workbook()
        sheet = workbook.active

@dataclass
class Results:
    url: str
    row: int = 0
    attempt_num: int = 0
    round_scores: list[int] = field(default_factory=list)
    round_distances: list[str] = field(default_factory=list)
    round_times: list[str] = field(default_factory=list)
    total_time : str = ""
    total_score: int = 0
    total_distance: str = ""

def find_next_empty_row(sheet, starting_row=2, col=2):
    """ Finds the next row for the times and scores to be added to """
    # Iterate over table to find next empty index
    next_row = starting_row
    for n, value in enumerate(sheet.iter_rows(min_row=starting_row,
                                 max_row=1048575,
                                 min_col=col,
                                 max_col=col,
                                 values_only=True)):
        if not value[0]:
            next_row += n
            break
        
    if next_row == starting_row:
        raise IndexError("All rows are full!")
    
    return next_row

def get_score_from_url(url):
    """ Scrapes info from geoguessr using selenium. """
    options = Options()
    driver=webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.get(url)

    # Find elements by class on the webpage
    driver.implicitly_wait(1)
    scores = driver.find_elements(By.CSS_SELECTOR, "[class*=results_score__]")
    details = driver.find_elements(By.CSS_SELECTOR, "[class*=results_scoreDetails__]")
    game_title = driver.find_element(By.CSS_SELECTOR, "[class*=info-card_title__]").text
    
    
    # Parse Text
    score_list = []
    dist_list = []
    time_list = []
    for score in scores:
        text = score.text
        text = text.replace(" pts", "")
        text = text.replace(",", "")
        score_list.append(int(text))

    for detail in details:
        dist, time = detail.text.split(" - ")
        time = time.replace(" min, ", ":")
        time = time.replace(" sec", "")
        try:
            if len(time.split(":")[1]) == 1:
                time = time.split(":")[0] + ":0" + time.split(":")[1]
        except IndexError:
            pass
        dist_list.append(dist)
        time_list.append(time)

    total_score = score_list.pop()
    total_dist = dist_list.pop()
    total_time = time_list.pop()

    return ((total_score, total_dist, total_time)), score_list, dist_list, time_list, game_title

def create_sheet(title):
    """ Create a new sheet with title that contains the necessary headings """
    workbook.create_sheet(title)
    sheet = workbook[title]

    # Set the width of the columns
    for n in range(66, 82): #                       "I"                "Q"
        sheet.column_dimensions[chr(n)].width = 0.5 if n == 72 else 55 if n == 81 else 14

    # Populate headings
    sheet["B2"] = "Attempt #"
    for n in range(1, 6):
        sheet[f"{chr(66 + n)}2"] = f"Round {n} Time"
        sheet[f"{chr(72 + n)}2"] = f"Round {n} Score"
    sheet["N2"] = "Total Score"
    sheet["O2"] = "Total Time"
    sheet["P2"] = "Vs. Average"
    sheet["Q2"] = "Link"
    sheet["S3"] = "=AVERAGE(O:O)"
    sheet["S3"].number_format = u"mm:ss"
    
    return sheet

def add_score_to_sheet(results, game_title):
    """ Add score data to the {game_title} spreadsheet """
    # Max title length in excel is 31
    if len(game_title) > 31:
        game_title = game_title[:30]
        print(f"[WARNING] Title too long! Title has been shortened to {game_title}")

    # Switch to appropriate sheet
    if game_title in workbook.sheetnames:
        sheet = workbook[game_title]

    # Create sheet if one doesn't exist
    else:
        print(f"\n[WARNING] No sheet with title {game_title}. Creating...")
        sheet = create_sheet(game_title)
        
    ### Find next empty row in sheet and add the data in ###
    results.row = find_next_empty_row(sheet)
    results.attempt_num = results.row - 2

    # Attempt Number
    sheet[f"B{results.row}"] = results.attempt_num
    
    # Times & Scores
    for i, n in enumerate(range(67, 72)):
        try:
            minute, second = results.round_times[i].split(":")
        except ValueError:
            minute = 0; second = results.round_times[i]
        time = datetime.time(minute=int(minute), second=int(second))
        
        sheet[f"{chr(n)}{results.row}"] = time
        sheet[f"{chr(n)}{results.row}"].number_format = u"mm:ss"
        sheet[f"{chr(n+6)}{results.row}"] = results.round_scores[i]

    # Totals, vs average and link columns
    sheet[f"N{results.row}"] = f"=SUM(I{results.row}:M{results.row})"
    sheet[f"O{results.row}"] = f"=SUM(C{results.row}:G{results.row})"
    sheet[f"P{results.row}"] = f"=O{results.row}-$S$3"
    sheet[f"Q{results.row}"] = results.url

    sheet[f"O{results.row}"].number_format = u"mm:ss"
    sheet[f"P{results.row}"].number_format = u"mm:ss"

    # There is negative time on the spreadsheet so 1904 date system used
    workbook.epoch = openpyxl.utils.datetime.CALENDAR_MAC_1904
    workbook.save(filename=FILENAME)
    print("[INFO] Workbook saved!")

def main():
    url = input("Please enter the results URL (type quit to exit): ")
    while url != "quit":
        results = Results(url)
        totals, scores, dists, times, game_title = get_score_from_url(url)
        results.total_score = totals[0]
        results.total_distance = totals[1]
        results.total_time = totals[2]
        results.round_scores = scores
        results.round_distances = dists
        results.round_times = times
        add_score_to_sheet(results, game_title)
            
        url = input("Please enter the results URL (type quit to exit): ")

if __name__ == "__main__":
    main()
